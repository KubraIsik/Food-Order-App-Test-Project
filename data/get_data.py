import openpyxl

xlsxFile = openpyxl.load_workbook("data/data.xlsx")

def get_restaurant_paths():
    sheet = xlsxFile["restaurant_paths"]
    rows = sheet.max_row
    data = {}
    for i in range(2,rows + 1):
        restaurant_id = sheet.cell(i,1).value 
        restaurant_path = sheet.cell(i,2).value 
        #data.append(restaurant_path)
        data[restaurant_id] =  restaurant_path # store as restaurant_id : restaurant_path pairs
    return data

def get_search_items(item_file_name):
    sheet = xlsxFile[item_file_name]
    rows = sheet.max_row
    data = []
    for i in range(2,rows + 1):
        search_item_word = sheet.cell(i,1).value 
        data.append(search_item_word)
    return data
